import io
import logging
from datetime import datetime, timedelta, timezone
from typing import List
from urllib.parse import quote

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from config import get_x86_specs_set, get_aarch64_specs_set, get_ies_specs_set
from database import get_db
from models import EcsServer, EcsConfig, EvsVolume, CesMetricDataDay, CesMetricData
from schemas import (
    CloudVmListRequest, CloudVmItem, ExportReportRequest,
    ApiResponse, PageData, ServerDetailItem, MetricDataPoint,
)
from services.ces_service import CesService

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/cloud-vm", tags=["cloud-vm"])


def _parse_safe_int(value: str) -> int:
    if not value or not value.strip():
        return 0
    try:
        return int(value.strip())
    except ValueError:
        return 0


def _map_status(status: str) -> str:
    if not status:
        return "未知"
    lower = status.lower()
    if lower == "active":
        return "运行中"
    if lower in ("stopped", "shutoff"):
        return "已关机"
    if lower == "error":
        return "异常"
    return status


def _determine_cpu_arch(flavor_name: str) -> str:
    if not flavor_name:
        return ""
    if flavor_name in get_x86_specs_set():
        return "X86"
    if flavor_name in get_aarch64_specs_set():
        return "arm"
    if flavor_name in get_ies_specs_set():
        return "IES"
    return ""


def _round2(val: float) -> float:
    return round(val * 100.0) / 100.0


def _filter_servers(servers: list, req) -> list:
    filtered = []
    for s in servers:
        if req.department and (not s.department or req.department not in s.department):
            continue
        if req.appSystem and (not s.app_system or req.appSystem not in s.app_system):
            continue
        if req.hostName and (not s.name or req.hostName not in s.name):
            continue
        if req.ipAddress and (not s.ip_address or req.ipAddress not in s.ip_address):
            continue
        if req.status and _map_status(s.status) != req.status:
            continue
        if req.cpu and (not s.flavor_vcpus or s.flavor_vcpus != req.cpu):
            continue
        if req.memory and (not s.flavor_ram or s.flavor_ram != req.memory):
            continue
        if req.systemDiskMin is not None and (not s.system_disk or _parse_safe_int(s.system_disk) < req.systemDiskMin):
            continue
        if req.systemDiskMax is not None and (not s.system_disk or _parse_safe_int(s.system_disk) > req.systemDiskMax):
            continue
        if req.dataDiskMin is not None and (not s.data_disk or _parse_safe_int(s.data_disk) < req.dataDiskMin):
            continue
        if req.dataDiskMax is not None and (not s.data_disk or _parse_safe_int(s.data_disk) > req.dataDiskMax):
            continue
        filtered.append(s)
    return filtered


def _to_cloud_vm_item(server: EcsServer) -> CloudVmItem:
    return CloudVmItem(
        id=server.server_id,
        hostName=server.name,
        department=server.department,
        appSystem=server.app_system,
        ipAddress=server.ip_address if server.ip_address else server.access_ipv4,
        status=_map_status(server.status),
        os=server.os_type,
        spec=server.flavor_name,
        architecture=_determine_cpu_arch(server.flavor_name),
        region=server.availability_zone,
        cpu=_parse_safe_int(server.flavor_vcpus),
        memory=_parse_safe_int(server.flavor_ram) // 1024,
        systemDisk=_parse_safe_int(server.system_disk),
        dataDisk=_parse_safe_int(server.data_disk),
    )


def _calc_metric_stats(
    host_metrics: List[CesMetricDataDay],
    primary_max, primary_avg, primary_min,
    secondary_max, secondary_avg, secondary_min,
) -> list:
    if not host_metrics:
        return [0, 0, 0]

    max_vals = []
    avg_vals = []
    min_vals = []

    for d in host_metrics:
        max_val = primary_max(d) if primary_max(d) is not None and primary_max(d) != 0 else secondary_max(d)
        avg_val = primary_avg(d) if primary_avg(d) is not None and primary_avg(d) != 0 else secondary_avg(d)
        min_val = primary_min(d) if primary_min(d) is not None and primary_min(d) != 0 else secondary_min(d)

        if max_val is not None: max_vals.append(max_val)
        if avg_val is not None: avg_vals.append(avg_val)
        if min_val is not None: min_vals.append(min_val)

    return [
        max(max_vals) if max_vals else 0,
        sum(avg_vals) / len(avg_vals) if avg_vals else 0,
        min(min_vals) if min_vals else 0,
    ]


@router.post("/list")
def list_cloud_vms(request: CloudVmListRequest, db: Session = Depends(get_db)):
    all_servers = db.query(EcsServer).all()
    filtered = _filter_servers(all_servers, request)
    items = [_to_cloud_vm_item(s) for s in filtered]

    page_num = request.pageNum or 1
    page_size = request.pageSize or 10
    total = len(items)
    from_index = (page_num - 1) * page_size
    to_index = min(from_index + page_size, total)

    page_list = items[from_index:to_index] if from_index < total else []

    page_data = PageData(list=page_list, total=total, pageNum=page_num, pageSize=page_size)
    return ApiResponse.success(page_data)


@router.post("/export")
def export_report(request: ExportReportRequest, db: Session = Depends(get_db)):
    start_time = request.startDate + " 00:00:00"
    end_time = request.endDate + " 23:59:59"

    all_servers = db.query(EcsServer).all()
    servers = _filter_servers(all_servers, request)

    config_map = {}
    for c in db.query(EcsConfig).all():
        if c.project_id:
            config_map[c.project_id] = c

    all_volumes = db.query(EvsVolume).all()
    volumes_by_server = {}
    for v in all_volumes:
        if v.server_id:
            volumes_by_server.setdefault(v.server_id, []).append(v)

    server_ids = [s.server_id for s in servers]
    metric_map = {}
    if server_ids:
        for m in db.query(CesMetricDataDay).filter(
            CesMetricDataDay.instance_id.in_(server_ids),
            CesMetricDataDay.timestamp >= start_time,
            CesMetricDataDay.timestamp <= end_time,
        ).all():
            metric_map.setdefault(m.instance_id, []).append(m)

    wb = Workbook()
    sheet = wb.active
    sheet.title = "弹性云服务器报告"

    headers = [
        "弹性云服务器ID", "弹性云服务器名称", "弹性云服务器创建时间", "运行状态",
        "所属部门", "所属应用", "所属区域", "所属网络分区",
        "操作系统类型", "镜像ID", "镜像名称", "云主机规格", "CPU架构",
        "CPU(核)", "内存(GB)", "系统盘(GB)", "数据盘(GB)",
        "高IO数据盘(GB)", "超高IO数据盘(GB)", "挂载的云硬盘",
        "项目ID", "domainID", "IPV4地址", "IPV6地址", "region", "弹性公网IP",
        "CPU使用率峰值(%)", "CPU使用率均值(%)", "CPU使用率最小值(%)",
        "内存使用率峰值(%)", "内存使用率均值(%)", "内存使用率最小值(%)",
        "磁盘使用率峰值(%)", "磁盘使用率均值(%)", "磁盘使用率最小值(%)",
    ]

    for col_idx, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_idx, value=header)

    for row_idx, server in enumerate(servers, 2):
        config = config_map.get(server.project_id)
        region_name = config.region_name if config else ""
        network_zone = config.network_zone if config and config.network_zone else ""
        endpoint = config.endpoint if config else ""

        server_volumes = volumes_by_server.get(server.server_id, [])

        high_io_data_disk = 0
        ultra_io_data_disk = 0
        disk_desc_list = []
        for vol in server_volumes:
            size = vol.size if vol.size else 0
            is_sys = vol.bootable is not None and vol.bootable.lower() == "true"
            type_label = vol.volume_type
            if type_label in ("SSD", "ESSD", "ESSD2"):
                io_label = "超高IO"
                if not is_sys: ultra_io_data_disk += size
            elif type_label in ("SAS", "GPSSD", "GPSSD2"):
                io_label = "高IO"
                if not is_sys: high_io_data_disk += size
            else:
                io_label = type_label or "普通IO"
                if not is_sys: high_io_data_disk += size
            disk_type = "系统盘" if is_sys else "数据盘"
            disk_desc_list.append(f"{size}GB（{io_label}、{disk_type}）")

        cpu_arch = _determine_cpu_arch(server.flavor_name)

        host_metrics = metric_map.get(server.server_id, [])

        row_data = [
            server.server_id or "",
            server.name or "",
            server.created_at or "",
            _map_status(server.status),
            server.department or "",
            server.app_system or "",
            region_name,
            network_zone,
            server.os_type or "",
            server.image_id or "",
            server.image_name or "",
            server.flavor_name or "",
            cpu_arch,
            _parse_safe_int(server.flavor_vcpus),
            _parse_safe_int(server.flavor_ram) // 1024,
            _parse_safe_int(server.system_disk),
            _parse_safe_int(server.data_disk),
            high_io_data_disk,
            ultra_io_data_disk,
            "\n".join(disk_desc_list),
            server.project_id or "",
            server.tenant_id or "",
            server.ip_address or "",
            server.ipv6 or "",
            endpoint,
            server.public_eip or "",
        ]

        cpu_stats = _calc_metric_stats(host_metrics,
            lambda d: d.cpu_util_max, lambda d: d.cpu_util_avg, lambda d: d.cpu_util_min,
            lambda d: d.cpu_usage_max, lambda d: d.cpu_usage_avg, lambda d: d.cpu_usage_min)
        row_data.extend([_round2(v) for v in cpu_stats])

        mem_stats = _calc_metric_stats(host_metrics,
            lambda d: d.mem_util_max, lambda d: d.mem_util_avg, lambda d: d.mem_util_min,
            lambda d: d.mem_used_percent_max, lambda d: d.mem_used_percent_avg, lambda d: d.mem_used_percent_min)
        row_data.extend([_round2(v) for v in mem_stats])

        disk_stats = _calc_metric_stats(host_metrics,
            lambda d: d.disk_util_inband_max, lambda d: d.disk_util_inband_avg, lambda d: d.disk_util_inband_min,
            lambda d: d.disk_used_percent_max, lambda d: d.disk_used_percent_avg, lambda d: d.disk_used_percent_min)
        row_data.extend([_round2(v) for v in disk_stats])

        for col_idx, val in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, len(headers) + 1):
        sheet.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A" + chr(64 + col_idx - 26)].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = f"弹性云服务器报告_{request.startDate}_{request.endDate}.xlsx"
    encoded_name = quote(file_name)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"},
    )


@router.get("/fetch-metric-data")
def fetch_metric_data(
    startDate: str = Query(...),
    endDate: str = Query(...),
    serverId: str = Query(None),
    db: Session = Depends(get_db),
):
    try:
        fmt = "%Y-%m-%d %H:%M:%S"
        from_dt = datetime.strptime(startDate + " 00:00:00", fmt)
        to_dt = datetime.strptime(endDate + " 00:00:00", fmt)
        from_ts = int(from_dt.timestamp() * 1000)
        to_ts = int(to_dt.timestamp() * 1000)

        ces_service = CesService(db)
        if serverId:
            ces_service.fetch_and_save_metric_data(from_ts, to_ts, serverId)
        else:
            ces_service.fetch_and_save_metric_data(from_ts, to_ts)

        return ApiResponse.success("监控数据采集成功")
    except Exception as e:
        return ApiResponse.error(500, f"监控数据采集失败: {str(e)}")


@router.get("/detail/{server_id}")
def get_server_detail(server_id: str, db: Session = Depends(get_db)):
    server = db.query(EcsServer).filter_by(server_id=server_id).first()
    if not server:
        return ApiResponse.error(404, "云主机不存在")

    config = db.query(EcsConfig).filter_by(project_id=server.project_id).first()
    region_name = config.region_name if config else ""
    network_zone = config.network_zone if config and config.network_zone else ""

    volumes = db.query(EvsVolume).filter_by(server_id=server_id).all()
    volume_list = []
    for vol in volumes:
        is_sys = vol.bootable is not None and vol.bootable.lower() == "true"
        io_label = "普通IO"
        if vol.volume_type in ("SSD", "ESSD", "ESSD2"):
            io_label = "超高IO"
        elif vol.volume_type in ("SAS", "GPSSD", "GPSSD2"):
            io_label = "高IO"
        volume_list.append({
            "volumeId": vol.volume_id,
            "name": vol.name or "",
            "size": vol.size or 0,
            "volumeType": vol.volume_type or "",
            "ioLabel": io_label,
            "bootable": is_sys,
            "diskType": "系统盘" if is_sys else "数据盘",
        })

    detail = ServerDetailItem(
        id=server.server_id,
        hostName=server.name,
        department=server.department,
        appSystem=server.app_system,
        ipAddress=server.ip_address if server.ip_address else server.access_ipv4,
        ipv6=server.ipv6,
        publicEip=server.public_eip,
        status=_map_status(server.status),
        os=server.os_type,
        osName=server.os_name,
        spec=server.flavor_name,
        architecture=_determine_cpu_arch(server.flavor_name),
        region=server.availability_zone,
        availabilityZone=server.availability_zone,
        cpu=_parse_safe_int(server.flavor_vcpus),
        memory=_parse_safe_int(server.flavor_ram) // 1024,
        systemDisk=_parse_safe_int(server.system_disk),
        dataDisk=_parse_safe_int(server.data_disk),
        createdAt=server.created_at,
        imageName=server.image_name,
        projectId=server.project_id,
        regionName=region_name,
        networkZone=network_zone,
        volumes=volume_list,
    )
    return ApiResponse.success(detail)


@router.get("/metric-data/{server_id}")
def get_server_metric_data(
    server_id: str,
    startDate: str = Query(...),
    endDate: str = Query(...),
    period: str = Query("5min"),
    db: Session = Depends(get_db),
):
    start_time = startDate + " 00:00:00"
    end_time = endDate + " 23:59:59"

    if period == "day":
        query = db.query(CesMetricDataDay).filter(
            CesMetricDataDay.instance_id == server_id,
            CesMetricDataDay.timestamp >= start_time,
            CesMetricDataDay.timestamp <= end_time,
        ).order_by(CesMetricDataDay.timestamp).all()
    else:
        query = db.query(CesMetricData).filter(
            CesMetricData.instance_id == server_id,
            CesMetricData.timestamp >= start_time,
            CesMetricData.timestamp <= end_time,
        ).order_by(CesMetricData.timestamp).all()

    data_points = []
    for d in query:
        cpu_max = d.cpu_util_max if d.cpu_util_max and d.cpu_util_max != 0 else (d.cpu_usage_max or 0)
        cpu_avg = d.cpu_util_avg if d.cpu_util_avg and d.cpu_util_avg != 0 else (d.cpu_usage_avg or 0)
        cpu_min = d.cpu_util_min if d.cpu_util_min and d.cpu_util_min != 0 else (d.cpu_usage_min or 0)
        mem_max = d.mem_util_max if d.mem_util_max and d.mem_util_max != 0 else (d.mem_used_percent_max or 0)
        mem_avg = d.mem_util_avg if d.mem_util_avg and d.mem_util_avg != 0 else (d.mem_used_percent_avg or 0)
        mem_min = d.mem_util_min if d.mem_util_min and d.mem_util_min != 0 else (d.mem_used_percent_min or 0)
        disk_max = d.disk_util_inband_max if d.disk_util_inband_max and d.disk_util_inband_max != 0 else (d.disk_used_percent_max or 0)
        disk_avg = d.disk_util_inband_avg if d.disk_util_inband_avg and d.disk_util_inband_avg != 0 else (d.disk_used_percent_avg or 0)
        disk_min = d.disk_util_inband_min if d.disk_util_inband_min and d.disk_util_inband_min != 0 else (d.disk_used_percent_min or 0)

        data_points.append(MetricDataPoint(
            timestamp=d.timestamp,
            cpuUtilMax=_round2(cpu_max),
            cpuUtilAvg=_round2(cpu_avg),
            cpuUtilMin=_round2(cpu_min),
            memUtilMax=_round2(mem_max),
            memUtilAvg=_round2(mem_avg),
            memUtilMin=_round2(mem_min),
            diskUtilMax=_round2(disk_max),
            diskUtilAvg=_round2(disk_avg),
            diskUtilMin=_round2(disk_min),
        ))

    return ApiResponse.success(data_points)


@router.get("/export-metric/{server_id}")
def export_server_metric(
    server_id: str,
    startDate: str = Query(...),
    endDate: str = Query(...),
    period: str = Query("5min"),
    db: Session = Depends(get_db),
):
    server = db.query(EcsServer).filter_by(server_id=server_id).first()
    if not server:
        return ApiResponse.error(404, "云主机不存在")

    start_time = startDate + " 00:00:00"
    end_time = endDate + " 23:59:59"

    if period == "day":
        query = db.query(CesMetricDataDay).filter(
            CesMetricDataDay.instance_id == server_id,
            CesMetricDataDay.timestamp >= start_time,
            CesMetricDataDay.timestamp <= end_time,
        ).order_by(CesMetricDataDay.timestamp).all()
    else:
        query = db.query(CesMetricData).filter(
            CesMetricData.instance_id == server_id,
            CesMetricData.timestamp >= start_time,
            CesMetricData.timestamp <= end_time,
        ).order_by(CesMetricData.timestamp).all()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "监控数据"

    host_name = server.name or server_id
    headers = [
        "时间", "CPU使用率峰值(%)", "CPU使用率均值(%)", "CPU使用率最小值(%)",
        "内存使用率峰值(%)", "内存使用率均值(%)", "内存使用率最小值(%)",
        "磁盘使用率峰值(%)", "磁盘使用率均值(%)", "磁盘使用率最小值(%)",
    ]

    for col_idx, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_idx, value=header)

    for row_idx, d in enumerate(query, 2):
        cpu_max = d.cpu_util_max if d.cpu_util_max and d.cpu_util_max != 0 else (d.cpu_usage_max or 0)
        cpu_avg = d.cpu_util_avg if d.cpu_util_avg and d.cpu_util_avg != 0 else (d.cpu_usage_avg or 0)
        cpu_min = d.cpu_util_min if d.cpu_util_min and d.cpu_util_min != 0 else (d.cpu_usage_min or 0)
        mem_max = d.mem_util_max if d.mem_util_max and d.mem_util_max != 0 else (d.mem_used_percent_max or 0)
        mem_avg = d.mem_util_avg if d.mem_util_avg and d.mem_util_avg != 0 else (d.mem_used_percent_avg or 0)
        mem_min = d.mem_util_min if d.mem_util_min and d.mem_util_min != 0 else (d.mem_used_percent_min or 0)
        disk_max = d.disk_util_inband_max if d.disk_util_inband_max and d.disk_util_inband_max != 0 else (d.disk_used_percent_max or 0)
        disk_avg = d.disk_util_inband_avg if d.disk_util_inband_avg and d.disk_util_inband_avg != 0 else (d.disk_used_percent_avg or 0)
        disk_min = d.disk_util_inband_min if d.disk_util_inband_min and d.disk_util_inband_min != 0 else (d.disk_used_percent_min or 0)

        row_data = [
            d.timestamp,
            _round2(cpu_max), _round2(cpu_avg), _round2(cpu_min),
            _round2(mem_max), _round2(mem_avg), _round2(mem_min),
            _round2(disk_max), _round2(disk_avg), _round2(disk_min),
        ]
        for col_idx, val in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, len(headers) + 1):
        col_letter = chr(64 + col_idx) if col_idx <= 26 else "A" + chr(64 + col_idx - 26)
        sheet.column_dimensions[col_letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = f"{host_name}_监控数据_{startDate}_{endDate}.xlsx"
    encoded_name = quote(file_name)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"},
    )
